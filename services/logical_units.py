"""Discover independently schedulable logical files inside large containers."""

import csv
import bz2
import io
import hashlib
import re
import tarfile
import zipfile
from pathlib import Path, PurePosixPath
from urllib.parse import quote

from services.unified_parser import _zip_central_directory_preflight


STRUCTURED_TEXT_EXTENSIONS = {".csv", ".tsv", ".jsonl"}
WORKBOOK_EXTENSIONS = {".xlsx", ".xlsm"}
ARCHIVE_EXTENSIONS = {".zip", ".tar", ".tar.gz", ".tgz", ".tar.bz2", ".tbz2", ".tar.xz", ".txz", ".bz2", ".rar", ".7z"}


def _safe_member_name(value):
    name = str(value or "").replace("\\", "/").lstrip("/")
    normalized = str(PurePosixPath(name))
    if not normalized or normalized == "." or normalized.startswith("../") or "/../" in normalized:
        return None
    return normalized


def _base_unit(container, logical_path, name, extension, size, kind, **extra):
    return {
        "id": hashlib.sha1(logical_path.encode("utf-8", errors="replace")).hexdigest()[:16],
        "path": logical_path,
        "name": name,
        "kind": "logical_file",
        "logical_unit": True,
        "logical_kind": kind,
        "container_path": container.get("path"),
        "container_name": container.get("name"),
        "container_size": int(container.get("size") or 0),
        "container_modified_at_ns": int(container.get("modified_at_ns") or 0),
        "container_device": container.get("device"),
        "container_inode": container.get("inode"),
        "extension": str(extension or "").lower(),
        "size": max(0, int(size or 0)),
        "modified_at": container.get("modified_at"),
        "modified_at_ns": int(container.get("modified_at_ns") or 0),
        "content_analysis_allowed": True,
        **extra,
    }


def archive_member_units(root, container, max_units=250000):
    path = Path(root) / str(container.get("path") or "")
    archive_name = str(path).lower()
    if not any(archive_name.endswith(ext) for ext in ARCHIVE_EXTENSIONS):
        return
    count = 0
    def emit(member, size, compressed_size=0, encrypted=False):
        nonlocal count
        member = _safe_member_name(member)
        if not member:
            return None
        if count >= max_units:
            raise ValueError("压缩包逻辑成员超过安全上限 {}，未建立不完整队列".format(max_units))
        count += 1
        extension = Path(member).suffix.lower()
        logical_path = "{}::{}".format(container["path"], member)
        return _base_unit(
            container, logical_path, Path(member).name, extension,
            int(size or 0), "archive_member", member_name=member,
            compressed_size=int(compressed_size or 0), encrypted=bool(encrypted),
            promotion_allowed=not bool(encrypted), archive_format=archive_name,
        )
    if archive_name.endswith(".zip"):
        preflight = _zip_central_directory_preflight(path)
        if not preflight.get("safe"):
            raise ValueError("ZIP中央目录预检失败：{}".format(preflight.get("reason") or "invalid"))
        observed = int(preflight.get("observed_entries") or 0)
        if observed > max_units:
            raise ValueError("压缩包逻辑成员超过安全上限 {}，未建立不完整队列".format(max_units))
        with zipfile.ZipFile(str(path)) as archive:
            if len(archive.filelist) != observed:
                raise ValueError("ZIP中央目录在预检后发生变化")
            for info in archive.filelist:
                if not info.is_dir():
                    unit = emit(info.filename, info.file_size, info.compress_size, info.flag_bits & 0x1)
                    if unit:
                        yield unit
        return
    if archive_name.endswith((".tar", ".tar.gz", ".tgz", ".tar.bz2", ".tbz2", ".tar.xz", ".txz")):
        with tarfile.open(str(path), mode="r:*") as archive:
            for info in archive:
                if info.isfile():
                    unit = emit(info.name, info.size)
                    if unit:
                        yield unit
        return
    if archive_name.endswith(".bz2"):
        unit = emit(Path(str(path)).stem, path.stat().st_size)
        if unit:
            unit["compressed_stream"] = True
            yield unit
        return
    if archive_name.endswith(".rar"):
        try:
            import rarfile
        except ImportError as exc:
            raise ValueError("RAR逻辑成员需要安装 rarfile 依赖") from exc
        with rarfile.RarFile(str(path)) as archive:
            for info in archive.infolist():
                if not info.is_dir():
                    unit = emit(info.filename, info.file_size, info.compress_size, getattr(info, "needs_password", lambda: False)())
                    if unit:
                        yield unit
        return
    try:
        import py7zr
    except ImportError as exc:
        raise ValueError("7z逻辑成员需要安装 py7zr 依赖") from exc
    with py7zr.SevenZipFile(str(path), mode="r") as archive:
        for info in archive.list():
            name = getattr(info, "filename", "")
            if not getattr(info, "is_directory", False):
                unit = emit(name, getattr(info, "uncompressed", 0))
                if unit:
                    yield unit


def structured_text_units(root, container, partition_bytes=1024 * 1024):
    path = Path(root) / str(container.get("path") or "")
    size = int(path.stat().st_size if path.exists() else container.get("size") or 0)
    extension = str(container.get("extension") or "").lower()
    if extension not in STRUCTURED_TEXT_EXTENSIONS or size <= partition_bytes:
        return
    if not path.exists():
        # Inventory-only callers may provide metadata without a local source;
        # retain deterministic queue accounting until the source is mounted.
        for index, start in enumerate(range(0, size, partition_bytes), 1):
            end = min(size, start + partition_bytes)
            yield _base_unit(
                container, "{}::partition/{:06d}".format(container["path"], index),
                "{}-part-{:06d}{}".format(Path(container["path"]).stem, index, extension),
                extension, end - start, "structured_text_partition",
                byte_start=start, byte_end=end, partition_index=index,
                include_header=False, record_boundary="unknown_source",
            )
        return
    include_header = extension in {".csv", ".tsv"}
    ranges = []
    with path.open("rb") as source:
        header = source.readline() if include_header else b""
        data_start = source.tell() if include_header else 0
        data_ranges = []
        chunk_start = data_start
        chunk_end = data_start
        for line in source:
            chunk_end = source.tell()
            if chunk_end - chunk_start > partition_bytes and chunk_end > chunk_start:
                data_ranges.append((chunk_start, chunk_end - len(line)))
                chunk_start = chunk_end - len(line)
        if chunk_end > chunk_start:
            data_ranges.append((chunk_start, chunk_end))
        if include_header and data_ranges:
            ranges.append((0, data_ranges[0][1]))
            ranges.extend(data_ranges[1:])
        elif include_header and header:
            ranges.append((0, len(header)))
        else:
            ranges = data_ranges
    if include_header and header:
        # Non-first CSV partitions carry a header explicitly; JSONL never does.
        pass
    index = 1
    for start, end in ranges:
        logical_path = "{}::partition/{:06d}".format(container["path"], index)
        actual_size = end - start + (len(header) if include_header and index > 1 else 0)
        yield _base_unit(
            container, logical_path,
            "{}-part-{:06d}{}".format(Path(container["path"]).stem, index, extension),
            extension, actual_size, "structured_text_partition",
            byte_start=start, byte_end=end, partition_index=index,
            include_header=bool(include_header and index > 1),
            record_boundary="line",
        )
        index += 1


def workbook_units(root, container, rows_per_partition=10000, max_units=250000):
    extension = str(container.get("extension") or "").lower()
    if extension not in WORKBOOK_EXTENSIONS:
        return
    try:
        import openpyxl
    except ImportError:
        return
    path = Path(root) / str(container.get("path") or "")
    workbook = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    try:
        count = 0
        for worksheet in workbook.worksheets:
            max_row = max(1, int(worksheet.max_row or 1))
            start_row = 1
            part = 1
            while start_row <= max_row:
                if count >= max_units:
                    raise ValueError("工作簿逻辑分片超过安全上限 {}".format(max_units))
                end_row = min(max_row, start_row + rows_per_partition - 1)
                measured = io.StringIO(newline="")
                writer = csv.writer(measured)
                actual_rows = 0
                for row in worksheet.iter_rows(min_row=start_row, max_row=end_row, values_only=True):
                    writer.writerow(["" if value is None else value for value in row])
                    actual_rows += 1
                actual_size = len(measured.getvalue().encode("utf-8-sig"))
                encoded_sheet = quote(str(worksheet.title), safe="")
                logical_path = "{}::sheet/{}/part/{:06d}".format(
                    container["path"], encoded_sheet, part
                )
                yield _base_unit(
                    container, logical_path,
                    "{}-{}-part-{:06d}.csv".format(
                        Path(container["path"]).stem,
                        re.sub(r"[^\w\u4e00-\u9fff.-]+", "-", str(worksheet.title))[:80],
                        part,
                    ),
                    ".csv", actual_size, "workbook_rows",
                    sheet_name=str(worksheet.title), row_start=start_row,
                    row_end=end_row, partition_index=part,
                    actual_row_count=actual_rows, estimated_size=False,
                )
                count += 1
                start_row = end_row + 1
                part += 1
    finally:
        workbook.close()


def iter_logical_units(root, physical_files, partition_bytes=1024 * 1024,
                       rows_per_partition=10000, max_units_per_container=250000):
    """Yield logical children; ordinary physical files remain unchanged."""
    for container in physical_files:
        path_name = str(container.get("path") or "").lower()
        extension = str(container.get("extension") or Path(path_name).suffix).lower()
        if extension in ARCHIVE_EXTENSIONS or any(path_name.endswith(ext) for ext in ARCHIVE_EXTENSIONS):
            yield from archive_member_units(root, container, max_units=max_units_per_container)
        elif extension in STRUCTURED_TEXT_EXTENSIONS:
            yield from structured_text_units(root, container, partition_bytes=partition_bytes)
        elif extension in WORKBOOK_EXTENSIONS:
            yield from workbook_units(
                root, container, rows_per_partition=rows_per_partition,
                max_units=max_units_per_container,
            )


def container_node_for(unit):
    """Reconstruct the verified physical inventory identity for one unit."""
    return {
        "path": unit.get("container_path"),
        "name": unit.get("container_name") or Path(str(unit.get("container_path") or "")).name,
        "size": int(unit.get("container_size") or 0),
        "modified_at_ns": int(unit.get("container_modified_at_ns") or 0),
        "modified_at": unit.get("modified_at"),
        "device": unit.get("container_device"),
        "inode": unit.get("container_inode"),
        "extension": Path(str(unit.get("container_path") or "")).suffix.lower(),
    }
