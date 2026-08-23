"""Bounded, dependency-light profiling for CSV/XLSX/JSON data files."""

import csv
import codecs
import json
import os
import re
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

STRUCTURED_EXTENSIONS = {".csv", ".tsv", ".json", ".jsonl", ".xlsx", ".xlsm"}
SENSITIVE_PATTERNS = {
    "email": re.compile(r"email|e-mail|邮箱|邮件", re.I),
    "phone": re.compile(r"phone|mobile|tel|电话|手机", re.I),
    "id": re.compile(r"(^|[_ -])(id|身份证|证件|账号|账户|编号)($|[_ -])", re.I),
    "name": re.compile(r"name|姓名|联系人|客户名称", re.I),
    "address": re.compile(r"address|地址|住址", re.I),
}

ENTITY_PATTERNS = {
    "person": re.compile(r"name|姓名|联系人|负责人|作者|人员|员工|客户", re.I),
    "location": re.compile(r"address|地址|地点|位置|区域|城市|省|市|县|国家|地区|经纬度", re.I),
    "event": re.compile(r"event|事件|活动|事项|类型|状态|category|类别|主题", re.I),
}


def _env_int(name, default):
    import os
    try:
        return max(1, int(os.getenv(name, str(default))))
    except (TypeError, ValueError):
        return default


def _missing(value):
    return value is None or str(value).strip().lower() in {"", "null", "none", "nan", "n/a", "na", "-"}


def _number(value):
    if isinstance(value, bool) or value is None:
        return None
    try:
        text = str(value).strip().replace(",", "")
        return float(text) if text else None
    except (TypeError, ValueError):
        return None


def _date(value):
    if isinstance(value, datetime):
        return value.isoformat()
    text = str(value or "").strip()
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S"):
        try:
            return datetime.strptime(text[:19], fmt).isoformat()
        except ValueError:
            pass
    return None


def _kind(value):
    if _missing(value):
        return "missing"
    if _number(value) is not None:
        return "number"
    if _date(value):
        return "datetime"
    return "text"


def _profile_rows(rows, columns, row_count, duplicate_rows, max_values=20, limits=None, status="completed"):
    result = {}
    for name, values in columns.items():
        kinds = Counter(_kind(value) for value in values)
        numbers = [n for n in (_number(value) for value in values) if n is not None]
        dates = [d for d in (_date(value) for value in values) if d]
        non_missing = [str(value).strip() for value in values if not _missing(value)]
        missing_count = sum(1 for value in values if _missing(value))
        item = {
            "name": name,
            "inferred_type": kinds.most_common(1)[0][0] if kinds else "unknown",
            "missing_count": missing_count,
            "missing_ratio": round(missing_count / float(max(1, row_count)), 6),
            "unique_count": len(set(non_missing)),
            "sample_values": list(dict.fromkeys(non_missing))[:max_values],
            "type_counts": dict(kinds),
        }
        if numbers:
            ordered = sorted(numbers)
            item.update({"min": min(numbers), "max": max(numbers), "sum": round(sum(numbers), 6), "count": len(numbers), "mean": round(sum(numbers) / len(numbers), 6), "median": ordered[len(ordered) // 2]})
            if len(ordered) >= 4:
                q1, q3 = ordered[len(ordered) // 4], ordered[(len(ordered) * 3) // 4]
                iqr = q3 - q1
                item["outlier_count"] = sum(1 for value in numbers if value < q1 - 1.5 * iqr or value > q3 + 1.5 * iqr)
            if len(ordered) >= 5:
                buckets = Counter()
                lower, upper = min(numbers), max(numbers)
                width = (upper - lower) / 5.0 if upper != lower else 1.0
                for value in numbers:
                    bucket = min(4, int((value - lower) / width)) if width else 0
                    buckets["{}-{}".format(bucket + 1, 5)] += 1
                item["distribution"] = dict(buckets)
        if dates:
            item.update({"min_datetime": min(dates), "max_datetime": max(dates)})
        non_missing_counter = Counter(non_missing)
        if item["inferred_type"] == "text":
            item["top_values"] = [{"value": value, "count": count} for value, count in non_missing_counter.most_common(max_values)]
        sensitive = [label for label, pattern in SENSITIVE_PATTERNS.items() if pattern.search(str(name))]
        if sensitive:
            item["sensitive_categories"] = sensitive
        entities = [label for label, pattern in ENTITY_PATTERNS.items() if pattern.search(str(name))]
        if entities:
            item["entity_categories"] = entities
        result[str(name)] = item
    missing = sum(item["missing_count"] for item in result.values())
    total = max(1, row_count * len(result))
    completeness = max(0.0, 1.0 - missing / float(total))
    uniqueness = sum(1 for item in result.values() if item.get("unique_count", 0) >= max(1, row_count * 0.9)) / float(len(result) or 1)
    duplicate_penalty = min(0.25, duplicate_rows / float(max(1, row_count)))
    quality = round(max(0.0, min(100.0, (completeness * 70 + uniqueness * 30) * (1 - duplicate_penalty))), 2)
    if not row_count or not result:
        # An empty/invalid source is not a moderately healthy dataset merely
        # because there are no missing cells to count.
        quality = 0.0
    missing_columns = [name for name, item in result.items() if item["missing_count"]]
    numeric_columns = [name for name, item in result.items() if item["inferred_type"] == "number"]
    temporal_columns = [name for name, item in result.items() if item["inferred_type"] == "datetime"]
    sensitive_columns = [name for name, item in result.items() if item.get("sensitive_categories")]
    entity_columns = {
        category: [name for name, item in result.items() if category in (item.get("entity_categories") or [])]
        for category in ENTITY_PATTERNS
    }
    entity_columns = {key: value for key, value in entity_columns.items() if value}
    entity_statistics = {}
    for category, names in entity_columns.items():
        values = []
        for name in names:
            values.extend(str(value).strip() for value in columns.get(name, []) if not _missing(value))
        counts = Counter(values)
        entity_statistics[category] = {
            "columns": names,
            "distinct_count": len(counts),
            "top_values": [{"value": value, "count": count} for value, count in counts.most_common(20)],
        }
    recommendation_questions = []
    if temporal_columns:
        recommendation_questions.append("按时间字段统计记录量和关键数值的变化趋势")
    if numeric_columns:
        recommendation_questions.append("哪些数值字段的均值、最大值和异常值最值得关注？")
    if entity_statistics.get("person"):
        recommendation_questions.append("不同人物/责任人对应的记录量和数值差异是什么？")
    if entity_statistics.get("location"):
        recommendation_questions.append("不同地区/地点的数据分布和质量差异是什么？")
    if entity_statistics.get("event"):
        recommendation_questions.append("不同事件/类型/状态的数量和关键指标如何比较？")
    profile_limits = {
        "sampled": True,
        "max_rows": _env_int("MAX_STRUCTURED_PROFILE_ROWS", 100000),
    }
    profile_limits.update(limits or {})
    truncated = bool(profile_limits.get("truncated"))
    return {
        "schema_version": "structured-profile/1.1", "status": status, "row_count": row_count,
        "column_count": len(result), "columns": result, "duplicate_row_count": duplicate_rows,
        "quality_score": quality, "missing_columns": missing_columns, "numeric_columns": numeric_columns,
        "temporal_columns": temporal_columns, "sensitive_columns": sensitive_columns,
        "entity_columns": entity_columns, "entity_statistics": entity_statistics,
        "recommendation_questions": recommendation_questions[:8],
        "coverage": {
            "complete": bool(status == "completed" and not truncated),
            "truncated": truncated,
            "sampled_rows": row_count,
            "truncation_reasons": list(profile_limits.get("truncation_reasons") or []),
        },
        "value_judgment": {
            "usable": bool(row_count and result),
            "value_level": "高" if quality >= 80 and row_count >= 10 else ("中" if quality >= 50 else "待治理"),
            "reason": "包含可统计字段、{} 行记录，数据质量评分 {} / 100。".format(row_count, quality),
            "recommended_next_steps": (["脱敏后进行精确统计"] if sensitive_columns else []) + (["核查缺失值和重复记录"] if missing_columns or duplicate_rows else []) + (["按时间字段做趋势分析"] if temporal_columns else []),
        },
        "limits": profile_limits,
    }


def _detect_text_encoding(path, probe_bytes=1024 * 1024):
    """Detect the small set of encodings commonly emitted by Chinese tools.

    UTF-8 is tried first, then GB18030 (a superset that can decode GBK and
    GB2312).  The probe is bounded and never reads an entire large source.
    """
    with Path(path).open("rb") as stream:
        sample = stream.read(max(1, probe_bytes))
    if sample.startswith(b"\xef\xbb\xbf"):
        return "utf-8-sig"
    if sample.startswith((b"\xff\xfe", b"\xfe\xff", b"\xff\xfe\x00\x00", b"\x00\x00\xfe\xff")):
        return "utf-16"
    for encoding in ("utf-8", "gb18030"):
        try:
            sample.decode(encoding, errors="strict")
            return encoding
        except UnicodeDecodeError as exc:
            # A bounded probe can end halfway through a multibyte UTF-8
            # character.  Treat only that trailing fragment as incomplete;
            # invalid bytes in the middle still cause the next codec to be
            # tried.
            if encoding == "utf-8" and exc.start >= max(0, len(sample) - 4):
                try:
                    sample[: exc.start].decode("utf-8", errors="strict")
                    return encoding
                except UnicodeDecodeError:
                    pass
            continue
    # Keep the parser usable for mixed/broken exports.  Replacement characters
    # are counted and surfaced in the profile instead of being silently hidden.
    return "utf-8"


class _BoundedLineReader:
    """Yield decoded lines while reading at most ``limit`` source bytes."""

    def __init__(self, path, limit, encoding, max_line_bytes=None):
        self.stream = Path(path).open("rb")
        self.limit = max(1, int(limit))
        self.encoding = encoding
        self.max_line_bytes = max(1, int(max_line_bytes)) if max_line_bytes else None
        self.bytes_read = 0
        self.replacement_count = 0
        self.hit_limit = False
        self.cut_mid_record = False
        self.source_record_exceeded_limit = False
        self._discarding_line = False

    def __iter__(self):
        return self

    def __next__(self):
        while True:
            if self._discarding_line:
                # Consume the remainder of an overlong logical record in
                # bounded chunks.  It must not be exposed as a second record
                # to csv.reader/json.loads on the next iteration.
                while self.bytes_read < self.limit:
                    remaining = self.limit - self.bytes_read
                    discard_size = min(remaining, self.max_line_bytes or remaining)
                    discarded = self.stream.readline(discard_size)
                    if not discarded:
                        self._discarding_line = False
                        raise StopIteration
                    self.bytes_read += len(discarded)
                    if discarded.endswith((b"\n", b"\r")):
                        self._discarding_line = False
                        break
                if self._discarding_line or self.bytes_read >= self.limit:
                    self._discarding_line = False
                    self.hit_limit = self.bytes_read >= self.limit
                    raise StopIteration
                continue
            if self.bytes_read >= self.limit:
                self.hit_limit = True
                raise StopIteration
            remaining = self.limit - self.bytes_read
            read_size = remaining
            if self.max_line_bytes:
                read_size = min(read_size, self.max_line_bytes)
            raw = self.stream.readline(read_size)
            if not raw:
                self._discarding_line = False
                raise StopIteration
            self.bytes_read += len(raw)
            ended = raw.endswith((b"\n", b"\r"))
            if not ended and len(raw) >= read_size:
                self.cut_mid_record = True
                if self.max_line_bytes and read_size == self.max_line_bytes and self.bytes_read < self.limit:
                    self.source_record_exceeded_limit = True
                if self.bytes_read >= self.limit:
                    self.hit_limit = True
                # Yield the bounded prefix once, then consume the rest of this
                # overlong record without ever placing it in memory.
                self._discarding_line = True
            if self._discarding_line and ended:
                self._discarding_line = False
            text = raw.decode(self.encoding, errors="replace")
            self.replacement_count += text.count("\ufffd")
            if self._discarding_line and not ended:
                # This is the first prefix of an overlong line.  The next call
                # enters the discard branch below; the JSONL/CSV caller sees a
                # parse error/partial record rather than a fake second row.
                return text
            return text

    def close(self):
        self.stream.close()

    def __enter__(self):
        return self

    def __exit__(self, _type, _value, _traceback):
        self.close()


class _LimitedBinaryStream:
    """File-like adapter used by ijson so a malformed 5GB source is bounded."""

    def __init__(self, stream, limit):
        self.stream = stream
        self.limit = max(1, int(limit))
        self.bytes_read = 0

    def readable(self):
        return True

    def read(self, size=-1):
        remaining = self.limit - self.bytes_read
        if remaining <= 0:
            return b""
        if size is None or size < 0 or size > remaining:
            size = remaining
        data = self.stream.read(size)
        self.bytes_read += len(data)
        return data

    def readline(self, size=-1):
        remaining = self.limit - self.bytes_read
        if remaining <= 0:
            return b""
        if size is None or size < 0 or size > remaining:
            size = remaining
        data = self.stream.readline(size)
        self.bytes_read += len(data)
        return data


class _IncrementalJsonReader:
    """Incrementally decode JSON without materialising the source document."""

    def __init__(self, path, encoding, max_bytes, max_record_chars):
        self.source = Path(path).open("rb")
        self.stream = _LimitedBinaryStream(self.source, max_bytes)
        codec = "utf-8" if encoding == "utf-8-sig" else encoding
        self.decoder = codecs.getincrementaldecoder(codec)(errors="replace")
        self.buffer = ""
        self.eof = False
        self.max_record_chars = max(1024, int(max_record_chars))
        self.replacement_characters = 0

    @property
    def bytes_read(self):
        return self.stream.bytes_read

    def close(self):
        self.source.close()

    def _read_more(self):
        if self.eof:
            return False
        raw = self.stream.read(64 * 1024)
        if raw:
            text = self.decoder.decode(raw, final=False)
            self.replacement_characters += text.count("\ufffd")
            self.buffer += text
            return True
        tail = self.decoder.decode(b"", final=True)
        self.replacement_characters += tail.count("\ufffd")
        self.buffer += tail
        self.eof = True
        return bool(tail)

    def skip_whitespace(self):
        while True:
            self.buffer = self.buffer.lstrip()
            if self.buffer:
                return True
            if not self._read_more():
                return False

    def parse_value(self):
        decoder = json.JSONDecoder()
        while True:
            if not self.skip_whitespace():
                raise ValueError("JSON 文档意外结束")
            try:
                value, end = decoder.raw_decode(self.buffer)
                self.buffer = self.buffer[end:]
                return value
            except json.JSONDecodeError as exc:
                if self.eof:
                    raise ValueError("JSON 文档不完整：{}".format(exc)) from exc
                if len(self.buffer) > self.max_record_chars:
                    raise ValueError("单条 JSON 记录超过安全上限 {} 字符".format(self.max_record_chars)) from exc
                if not self._read_more():
                    raise ValueError("JSON 文档意外结束") from exc


def _iter_json_incremental(path, encoding, max_rows, max_bytes):
    """Parse a JSON array/root object with a bounded incremental reader.

    Returns ``None`` when the source cannot be classified from its first token;
    callers can then use the optional ijson path.  The returned values are
    already limited to dictionaries because those are the rows understood by
    the profile engine.
    """
    max_record_chars = _env_int("MAX_STRUCTURED_JSON_RECORD_CHARS", 16 * 1024 * 1024)
    reader = _IncrementalJsonReader(path, encoding, max_bytes, max_record_chars)
    rows = []
    stopped_for_row_limit = False
    metadata = {
        "streaming": True,
        "truncated": Path(path).stat().st_size > max_bytes,
        "truncation_reasons": ["byte_limit"] if Path(path).stat().st_size > max_bytes else [],
        "parse_errors": 0,
    }
    try:
        if not reader.skip_whitespace():
            return rows, metadata
        if reader.buffer.startswith("\ufeff"):
            reader.buffer = reader.buffer[1:]
            if not reader.skip_whitespace():
                return rows, metadata
        is_array = reader.buffer.startswith("[")
        if is_array:
            reader.buffer = reader.buffer[1:]
            while True:
                if not reader.skip_whitespace():
                    raise ValueError("JSON 数组缺少结束符")
                if reader.buffer.startswith("]"):
                    reader.buffer = reader.buffer[1:]
                    break
                if len(rows) >= max_rows:
                    metadata["truncated"] = True
                    metadata["truncation_reasons"].append("row_limit")
                    stopped_for_row_limit = True
                    break
                value = reader.parse_value()
                if isinstance(value, dict):
                    rows.append(value)
                else:
                    metadata["parse_errors"] += 1
                if not reader.skip_whitespace():
                    raise ValueError("JSON 数组缺少逗号或结束符")
                if reader.buffer.startswith(","):
                    reader.buffer = reader.buffer[1:]
                    continue
                if reader.buffer.startswith("]"):
                    reader.buffer = reader.buffer[1:]
                    break
                raise ValueError("JSON 数组成员之间缺少逗号")
        else:
            value = reader.parse_value()
            if isinstance(value, dict):
                rows.append(value)
            elif isinstance(value, list):
                rows = [item for item in value[:max_rows] if isinstance(item, dict)]
                if len(value) > max_rows:
                    metadata["truncated"] = True
                    metadata["truncation_reasons"].append("row_limit")
        if not stopped_for_row_limit:
            reader.skip_whitespace()
        if reader.buffer.strip() and not stopped_for_row_limit:
            # Trailing non-whitespace is malformed, even if the leading rows
            # were valid.  Keep the rows but make the partial result visible.
            raise ValueError("JSON 文档包含尾随非空内容")
    except ValueError as exc:
        metadata["parse_errors"] += 1
        metadata["parse_error"] = str(exc)[:300]
    finally:
        metadata["sampled_bytes"] = reader.bytes_read
        metadata["replacement_characters"] = reader.replacement_characters
        reader.close()
    return rows, metadata


def _column_names(header):
    seen, names = Counter(), []
    for index, name in enumerate(header or []):
        name = str(name).strip() or "column_{}".format(index + 1)
        seen[name] += 1
        names.append(name if seen[name] == 1 else "{}_{}".format(name, seen[name]))
    return names


def _iter_csv(path, max_rows, max_bytes):
    rows = []
    names = []
    encoding = _detect_text_encoding(path)
    meta = {
        "encoding": encoding,
        "source_bytes": Path(path).stat().st_size,
        "sampled_bytes": 0,
        "truncated": False,
        "truncation_reasons": [],
        "parse_errors": 0,
        "streaming": True,
    }
    with _BoundedLineReader(
        path, max_bytes, encoding,
        max_line_bytes=min(max_bytes, 64 * 1024 * 1024),
    ) as lines:
        reader = csv.reader(lines, delimiter="\t" if Path(path).suffix.lower() == ".tsv" else ",")
        try:
            header = next(reader, None)
            names = _column_names(header)
            if not names:
                meta["sampled_bytes"] = lines.bytes_read
                return rows, [], meta["sampled_bytes"], meta
            try:
                csv.field_size_limit(max(128 * 1024, min(max_bytes, 64 * 1024 * 1024)))
            except (OverflowError, ValueError):
                pass
            for values in reader:
                if len(rows) >= max_rows:
                    meta["truncated"] = True
                    meta["truncation_reasons"].append("row_limit")
                    break
                rows.append(dict(zip(names, values)))
        except csv.Error as exc:
            meta["parse_errors"] += 1
            meta["parse_error"] = str(exc)[:300]
        meta["sampled_bytes"] = lines.bytes_read
        meta["replacement_characters"] = lines.replacement_count
        if lines.hit_limit or meta["source_bytes"] > max_bytes:
            meta["truncated"] = True
            if "byte_limit" not in meta["truncation_reasons"]:
                meta["truncation_reasons"].append("byte_limit")
        if lines.cut_mid_record and "partial_record" not in meta["truncation_reasons"]:
            meta["truncation_reasons"].append("partial_record")
            if lines.source_record_exceeded_limit and "record_limit" not in meta["truncation_reasons"]:
                meta["truncation_reasons"].append("record_limit")
    return rows, names, meta["sampled_bytes"], meta


def _iter_json(path, max_rows, max_bytes):
    path = Path(path)
    source_bytes = path.stat().st_size
    encoding = _detect_text_encoding(path)
    meta = {
        "encoding": encoding,
        "source_bytes": source_bytes,
        "sampled_bytes": 0,
        "truncated": source_bytes > max_bytes,
        "truncation_reasons": ["byte_limit"] if source_bytes > max_bytes else [],
        "parse_errors": 0,
        "streaming": False,
    }
    rows = []

    if path.suffix.lower() == ".jsonl":
        meta["streaming"] = True
        record_limit = _env_int("MAX_STRUCTURED_JSON_RECORD_BYTES", 16 * 1024 * 1024)
        with _BoundedLineReader(path, max_bytes, encoding, max_line_bytes=record_limit) as lines:
            for line in lines:
                if not line.strip():
                    continue
                if len(rows) >= max_rows:
                    meta["truncated"] = True
                    meta["truncation_reasons"].append("row_limit")
                    break
                try:
                    value = json.loads(line)
                except (ValueError, TypeError) as exc:
                    meta["parse_errors"] += 1
                    meta.setdefault("parse_error_samples", []).append(str(exc)[:160])
                    continue
                if isinstance(value, dict):
                    rows.append(value)
                else:
                    meta["parse_errors"] += 1
            meta["sampled_bytes"] = lines.bytes_read
            meta["replacement_characters"] = lines.replacement_count
            if lines.hit_limit or meta["source_bytes"] > max_bytes or lines.cut_mid_record:
                meta["truncated"] = True
                if (lines.hit_limit or meta["source_bytes"] > max_bytes) and "byte_limit" not in meta["truncation_reasons"]:
                    meta["truncation_reasons"].append("byte_limit")
                if lines.cut_mid_record and "partial_record" not in meta["truncation_reasons"]:
                    meta["truncation_reasons"].append("partial_record")
                if lines.source_record_exceeded_limit and "record_limit" not in meta["truncation_reasons"]:
                    meta["truncation_reasons"].append("record_limit")
    else:
        # Keep one bounded implementation for every encoding.  Some streaming
        # JSON libraries materialise each object without a per-record limit;
        # the local reader enforces both a byte budget and a record budget even
        # when an input contains one pathological 5GB value.
        incremental_rows, incremental_meta = _iter_json_incremental(
            path, encoding, max_rows, max_bytes
        )
        rows = incremental_rows
        for key, value in incremental_meta.items():
            if key == "truncation_reasons":
                meta[key] = list(dict.fromkeys((meta.get(key) or []) + (value or [])))
            else:
                meta[key] = value
        meta["streaming"] = True

    if meta["source_bytes"] > max_bytes and "byte_limit" not in meta["truncation_reasons"]:
        meta["truncated"] = True
        meta["truncation_reasons"].append("byte_limit")
    names = list(dict.fromkeys(key for item in rows if isinstance(item, dict) for key in item))
    rows = [{name: item.get(name) for name in names} for item in rows if isinstance(item, dict)]
    return rows, names, meta["sampled_bytes"], meta


def profile_path(path, max_rows=None, max_bytes=None):
    path = Path(path)
    ext = path.suffix.lower()
    if ext not in STRUCTURED_EXTENSIONS:
        return None
    max_rows = max_rows or _env_int("MAX_STRUCTURED_PROFILE_ROWS", 100000)
    max_bytes = max_bytes or _env_int("MAX_STRUCTURED_PROFILE_BYTES", 256 * 1024 * 1024)
    if path.stat().st_size > max_bytes and ext in {".xlsx", ".xlsm"}:
        return {"schema_version": "structured-profile/1.1", "status": "skipped", "reason": "文件超过结构化画像大小上限", "limits": {"sampled": False, "source_bytes": path.stat().st_size, "max_bytes": max_bytes, "truncated": True, "truncation_reasons": ["byte_limit"]}, "coverage": {"complete": False, "truncated": True, "sampled_rows": 0, "truncation_reasons": ["byte_limit"]}, "value_judgment": {"usable": False, "value_level": "待治理"}}
    metadata = {"encoding": None, "source_bytes": path.stat().st_size, "sampled_bytes": 0, "truncated": False, "truncation_reasons": [], "parse_errors": 0}
    book = None
    try:
        if ext in {".csv", ".tsv"}:
            rows, names, consumed, metadata = _iter_csv(path, max_rows, max_bytes)
        elif ext in {".json", ".jsonl"}:
            rows, names, consumed, metadata = _iter_json(path, max_rows, max_bytes)
        else:
            from openpyxl import load_workbook
            book = load_workbook(str(path), read_only=True, data_only=True)
            rows = []
            names = []
            processed_sheets = []
            skipped_sheets = []
            sheet_row_counts = {}
            worksheets = list(book.worksheets)
            budget_exhausted = False
            for sheet_index, sheet in enumerate(worksheets):
                if budget_exhausted:
                    skipped_sheets.append({
                        "name": sheet.title,
                        "state": getattr(sheet, "sheet_state", "visible"),
                        "reason": "global_row_limit",
                    })
                    continue
                iterator = None
                try:
                    # Read at most the header, the remaining global budget and
                    # one sentinel row.  Exhausting this bounded generator is
                    # important: openpyxl closes the worksheet ZipExtFile only
                    # when iteration naturally reaches StopIteration.
                    remaining_budget = max(0, max_rows - len(rows))
                    iterator = sheet.iter_rows(
                        values_only=True,
                        max_row=remaining_budget + 2,
                    )
                    header = next(iterator, None)
                    sheet_names = _column_names(header)
                    processed = {
                        "name": sheet.title,
                        "state": getattr(sheet, "sheet_state", "visible"),
                        "column_count": len(sheet_names),
                        "row_count": 0,
                    }
                    processed_sheets.append(processed)
                    sheet_row_counts[sheet.title] = 0
                    for name in sheet_names:
                        if name not in names:
                            names.append(name)
                    if not sheet_names:
                        continue
                    for values in iterator:
                        if len(rows) >= max_rows:
                            metadata["truncated"] = True
                            if "row_limit" not in metadata["truncation_reasons"]:
                                metadata["truncation_reasons"].append("row_limit")
                            processed["truncated"] = True
                            budget_exhausted = True
                            # Keep advancing through the one-row sentinel so
                            # the bounded generator closes its archive member.
                            continue
                        rows.append(dict(zip(sheet_names, values)))
                        processed["row_count"] += 1
                        sheet_row_counts[sheet.title] += 1
                except Exception as exc:
                    metadata["parse_errors"] += 1
                    skipped_sheets.append({
                        "name": sheet.title,
                        "state": getattr(sheet, "sheet_state", "visible"),
                        "reason": "worksheet_read_error",
                        "error": str(exc)[:160],
                    })
                finally:
                    # read_only worksheets keep a ZipExtFile open while their
                    # row generator is alive.  Explicitly close it when a
                    # global budget stops iteration early (notably on Windows).
                    close_iterator = getattr(iterator, "close", None)
                    if callable(close_iterator):
                        close_iterator()
            metadata.update({
                "worksheet_count": len(worksheets),
                "processed_worksheet_count": len(processed_sheets),
                "skipped_worksheet_count": len(skipped_sheets),
                "processed_worksheets": processed_sheets,
                "skipped_worksheets": skipped_sheets,
                "worksheet_row_counts": sheet_row_counts,
                "global_row_budget": max_rows,
            })
            consumed = path.stat().st_size
    except Exception as exc:
        return {"schema_version": "structured-profile/1.1", "status": "failed", "error": str(exc)[:300], "coverage": {"complete": False, "truncated": False, "sampled_rows": 0, "truncation_reasons": []}, "value_judgment": {"usable": False, "value_level": "待治理"}}
    finally:
        if book is not None:
            book.close()
    columns = defaultdict(list)
    duplicate_counter = Counter()
    for row in rows:
        duplicate_counter[tuple(str(row.get(name, "")).strip() for name in names)] += 1
        for name in names:
            columns[name].append(row.get(name))
    duplicate_rows = sum(count - 1 for count in duplicate_counter.values() if count > 1)
    status = "completed"
    if metadata.get("truncated") or metadata.get("parse_errors"):
        status = "partial" if rows else "failed"
    if metadata.get("replacement_characters"):
        metadata.setdefault("warnings", []).append("文本编码存在无法解码的字符，已使用替换字符；请核对原始文件。")
    if metadata.get("truncated"):
        metadata.setdefault("warnings", []).append("画像仅覆盖文件的有界采样范围，未将其视为完整数据集。")
    profile = _profile_rows(rows, columns, len(rows), duplicate_rows, limits={
        "max_rows": max_rows,
        "max_bytes": max_bytes,
        "sampled_bytes": consumed,
        "source_bytes": path.stat().st_size,
        "encoding": metadata.get("encoding"),
        "truncated": bool(metadata.get("truncated")),
        "truncation_reasons": list(dict.fromkeys(metadata.get("truncation_reasons") or [])),
        "parse_errors": int(metadata.get("parse_errors") or 0),
        "streaming": bool(metadata.get("streaming", False)),
        "worksheet_count": metadata.get("worksheet_count"),
        "processed_worksheet_count": metadata.get("processed_worksheet_count"),
        "skipped_worksheet_count": metadata.get("skipped_worksheet_count"),
        "processed_worksheets": metadata.get("processed_worksheets"),
        "skipped_worksheets": metadata.get("skipped_worksheets"),
        "worksheet_row_counts": metadata.get("worksheet_row_counts"),
        "global_row_budget": metadata.get("global_row_budget"),
    }, status=status)
    profile["source"] = {"path": str(path), "extension": ext, "size": path.stat().st_size, "encoding": metadata.get("encoding")}
    if metadata.get("parse_error"):
        profile["warnings"] = ["结构化数据解析未完全成功：{}".format(metadata["parse_error"])]
    if metadata.get("warnings"):
        profile.setdefault("warnings", []).extend(metadata["warnings"])
    return profile
